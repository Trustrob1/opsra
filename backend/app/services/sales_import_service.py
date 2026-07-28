"""
app/services/sales_import_service.py
Sales bulk import service — GPM-1E (watermark update).

File format detection uses magic bytes (first 4 bytes), not extension:
  XLSX  — PK\\x03\\x04  → openpyxl
  XLS   — \\xd0\\xcf\\x11\\xe0 → xlrd  (pip install xlrd)
  CSV   — everything else → csv.reader

Watermark logic:
  - get_watermark(db, org_id, source_type, sheet_url)  → date | None
  - save_watermark(db, org_id, source_type, sheet_url, last_date)
  - validate_and_prepare_rows() flags rows at/before watermark as already_imported
  - Rows flagged already_imported are SEPARATE from duplicate_warnings
  - Both warnings are non-blocking — admin selects which rows to import via indices
"""
from __future__ import annotations

import csv
import io
import logging
import re
from datetime import date, datetime
from typing import Optional

import httpx

logger = logging.getLogger(__name__)

_XLSX_MAGIC = b'PK\x03\x04'
_XLS_MAGIC  = b'\xd0\xcf\x11\xe0'

CANONICAL_FIELDS = {
    "customer_name", "phone", "region", "amount",
    "sale_date", "channel", "source_team", "notes",
}

_ALIASES: dict[str, str] = {
    "customer_name": "customer_name", "customer": "customer_name",
    "name": "customer_name", "client": "customer_name", "client_name": "customer_name",
    "phone": "phone", "phone_number": "phone", "tel": "phone",
    "telephone": "phone", "mobile": "phone",
    "region": "region", "area": "region", "location": "region", "city": "region",
    "amount": "amount", "price": "amount", "sale_amount": "amount",
    "value": "amount", "revenue": "amount",
    "sale_date": "sale_date", "date": "sale_date", "sold_date": "sale_date",
    "transaction_date": "sale_date",
    "channel": "channel", "source": "channel", "sale_channel": "channel",
    "source_team": "source_team", "team": "source_team", "team_name": "source_team",
    "sales_team": "source_team",
    "notes": "notes", "note": "notes", "remarks": "notes",
    "comment": "notes", "comments": "notes",
}

_DATE_FORMATS = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _normalise_header(raw: str, aliases: Optional[dict] = None) -> Optional[str]:
    aliases = aliases if aliases is not None else _ALIASES
    key = raw.strip().lower().replace(" ", "_").replace("-", "_")
    return aliases.get(key)


# ---------------------------------------------------------------------------
# REPORTS-DEPT-1 Phase 4: daily-aggregate sales import
#
# Distinct from CANONICAL_FIELDS/_ALIASES above, which are for individual
# named-customer transactions. This is for a daily aggregate sheet — one
# row per day per rep, with separate per-product-line revenue columns
# (e.g. Mattress Revenue, Pillow Revenue) and no customer at all.
# ---------------------------------------------------------------------------

_AGGREGATE_ALIASES: dict[str, str] = {
    "date": "sale_date", "sale_date": "sale_date",
    "sales_rep": "rep_name", "rep": "rep_name",
    "salesperson": "rep_name", "rep_name": "rep_name",
    "mattress_revenue": "mattress_revenue",
    "pillow_revenue": "pillow_revenue",
}


def _aggregate_rows_to_dicts(rows: list) -> list[dict]:
    """
    Maps raw spreadsheet rows for a daily aggregate sales sheet into
    normalised dicts. Mirrors _rows_to_dicts()'s shape/behaviour but uses
    _AGGREGATE_ALIASES instead of the named-customer _ALIASES.
    """
    if not rows:
        raise ValueError("File is empty — no rows found.")
    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_map: dict[int, str] = {}
    for idx, raw_h in enumerate(raw_headers):
        canonical = _normalise_header(raw_h, _AGGREGATE_ALIASES)
        if canonical:
            col_map[idx] = canonical
    if not col_map:
        raise ValueError(
            "No recognised columns found. Expected at least one of: "
            + ", ".join(sorted(set(_AGGREGATE_ALIASES.values())))
        )
    result: list[dict] = []
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record: dict = {field: None for field in set(_AGGREGATE_ALIASES.values())}
        for idx, canonical in col_map.items():
            cell_val = row[idx] if idx < len(row) else None
            record[canonical] = str(cell_val).strip() if cell_val is not None else None
        result.append(record)
    return result


def _match_rep(rep_name: Optional[str], org_id: str, db) -> tuple:
    """
    Match a raw rep name (as typed in the sheet) against this org's users
    by full_name (case-insensitive). Returns (user_id, team) or
    (None, None) if no match — an unmatched row still imports, just
    without rep/team attribution.
    """
    if not rep_name:
        return None, None
    try:
        result = (
            db.table("users")
            .select("id, team, full_name")
            .eq("org_id", org_id)
            .ilike("full_name", rep_name.strip())
            .limit(1)
            .execute()
        )
        rows = result.data or []
        if rows:
            return rows[0]["id"], rows[0].get("team")
    except Exception as exc:
        logger.warning("_match_rep failed org=%s rep=%s: %s", org_id, rep_name, exc)
    return None, None


# ---------------------------------------------------------------------------
# REPORTS-DEPT-1 Phase 4b: transaction-level sales log (per-sale, not
# aggregate) — for a workbook with one tab per region (e.g. Lagos, Abuja),
# feeding the Sales Record and Commissions tabs. Mirrors the reference
# commission-dashboard tool's column expectations exactly.
# ---------------------------------------------------------------------------

_TXN_ALIASES: dict[str, str] = {
    "date": "sale_date",
    "sales_rep": "rep_name_raw", "rep": "rep_name_raw", "salesperson": "rep_name_raw",
    "customer_name": "customer_name", "customer": "customer_name",
    "items_purchased": "items", "items": "items", "item": "items",
    "model": "model",
    "variant": "variant", "size": "variant", "spec": "variant",
    "units": "units", "qty": "units", "quantity": "units",
    "amount": "amount", "sale_amount": "amount",
    "status": "reconciliation_status",
    "reconciliation": "reconciliation_status",
    "reconciled": "reconciliation_status",
}


def parse_multi_sheet_xlsx(file_bytes: bytes) -> dict:
    """
    Reads EVERY sheet in an xlsx workbook (unlike parse_excel_file, which
    only reads wb.active — the one open when the file was last saved).
    Returns {sheet_name: raw_rows} so a caller can process each region's
    tab (e.g. "Lagos", "Abuja") separately and tag rows accordingly.
    S14: raises ValueError on any parse failure.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not read workbook: {exc}") from exc

    result: dict = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            result[sheet_name] = rows
    if not result:
        raise ValueError("Workbook has no sheets with data.")
    return result


def _txn_rows_to_dicts(rows: list, region: str) -> list[dict]:
    """
    Maps raw rows from one sheet/tab into normalised transaction dicts,
    tagged with the region (sheet name) they came from.
    """
    if not rows:
        return []
    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_map: dict[int, str] = {}
    for idx, raw_h in enumerate(raw_headers):
        canonical = _normalise_header(raw_h, _TXN_ALIASES)
        if canonical:
            col_map[idx] = canonical
    if not col_map:
        return []
    result: list[dict] = []
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record: dict = {field: None for field in set(_TXN_ALIASES.values())}
        for idx, canonical in col_map.items():
            cell_val = row[idx] if idx < len(row) else None
            if isinstance(cell_val, (datetime, date)):
                # Real Excel date cell — format directly as YYYY-MM-DD rather
                # than stringifying to "YYYY-MM-DD HH:MM:SS" and forcing
                # _parse_date to re-parse a shape its format list never
                # expected. This was silently rejecting real sale dates.
                record[canonical] = cell_val.strftime("%Y-%m-%d")
            else:
                record[canonical] = str(cell_val).strip() if cell_val is not None else None
        record["region"] = region
        result.append(record)
    return result


def validate_and_prepare_transaction_sales_rows(
    rows_by_region: dict,
    org_id: str,
    db,
    import_source: str,
    watermark_date: Optional[str] = None,
) -> dict:
    """
    Validate and prepare individual-sale rows (one row = one sale, not one
    day's total) for insertion into direct_sales, across multiple regions
    (one per sheet tab). Feeds the Sales Record and Commissions tabs.

    Matches the reference commission-dashboard tool's own rules:
      - customer_name and rep are the meaningful identity of a row (a row
        with neither is silently skipped, not an error — matches the
        reference tool's own `r => r[keys.rep] || r[keys.customer]` filter)
      - amount defaults to 0 if missing/unparseable (not an error) — the
        reference tool does the same (matchRate() * units still yields a
        meaningful commission even with a blank Amount column)
      - reconciliation_status: any value starting with "rec" (case-
        insensitive) -> "Reconciled", everything else -> "Pending",
        matching the reference tool's exact rule. Missing Status column
        entirely -> every row defaults to "Pending".
      - model falls back to the Items Purchased column if Model is blank,
        matching the reference tool's `row.model || row.items`.
    """
    from datetime import timezone
    now_iso = datetime.now(timezone.utc).isoformat()

    valid_rows:         list[dict] = []
    error_rows:         list[dict] = []
    duplicate_warnings: list[dict] = []
    already_imported:   list[dict] = []

    wm_date: Optional[date] = None
    if watermark_date:
        try:
            wm_date = datetime.strptime(watermark_date[:10], "%Y-%m-%d").date()
        except Exception:
            pass

    existing_keys: set = set()
    try:
        res = (
            db.table("direct_sales")
            .select("rep_name,sale_date,customer_name,model,amount")
            .eq("org_id", org_id)
            .eq("import_source", import_source)
            .execute()
        )
        for r in (res.data or []):
            if r.get("sale_date"):
                existing_keys.add((
                    (r.get("rep_name") or "").strip().lower(),
                    str(r["sale_date"])[:10],
                    (r.get("customer_name") or "").strip().lower(),
                    (r.get("model") or "").strip().lower(),
                    float(r["amount"]) if r.get("amount") is not None else None,
                ))
    except Exception:
        logger.warning("Phase 4c txn import: duplicate check query failed — skipping.")

    rep_cache: dict = {}

    def _get_rep(rep_name: Optional[str]):
        if not rep_name:
            return None, None
        key = rep_name.strip().lower()
        if key not in rep_cache:
            rep_cache[key] = _match_rep(rep_name, org_id, db)
        return rep_cache[key]

    model = (row.get("model") or "").strip() or (row.get("items") or "").strip() or None

            status_raw = (row.get("reconciliation_status") or "").strip().lower()
            reconciliation_status = "Reconciled" if status_raw.startswith("rec") else "Pending"

            rep_id, rep_team = _get_rep(rep_name_raw)

            if wm_date:
                try:
                    row_date = datetime.strptime(sale_date, "%Y-%m-%d").date()
                    if row_date <= wm_date:
                        already_imported.append({"row": row_num, "region": region, "sale_date": sale_date})
                except Exception:
                    pass

            dup_key = (rep_id, sale_date, (customer_name or "").lower(), (model or "").lower(), amount)
            if dup_key in existing_keys:
                duplicate_warnings.append({"row": row_num, "region": region, "sale_date": sale_date, "customer_name": customer_name})

            notes = None
            if rep_name_raw and not rep_id:
                notes = f"Unmatched rep on sheet: {rep_name_raw}"

            valid_rows.append({
                "org_id":                org_id,
                "customer_name":         customer_name,
                "amount":                amount,
                "currency":              "NGN",
                "sale_date":             sale_date,
                "channel":               "other",
                "region":                region,
                "model":                 model,
                "units":                 units,
                "reconciliation_status": reconciliation_status,
                "source_team":           rep_team,
                "recorded_by":           rep_id,
                "notes":                 notes,
                "import_source":         import_source,
                "created_at":            now_iso,
                "updated_at":            now_iso,
            })

            raw_date = str(row.get("sale_date") or "").strip()
            sale_date = _parse_date(raw_date)
            if not sale_date:
                error_rows.append({
                    "row": row_num, "region": region,
                    "message": f"Invalid date: '{raw_date}'. Use YYYY-MM-DD, DD/MM/YYYY, or MM/DD/YYYY",
                })
                continue

            try:
                units = float(str(row.get("units") or "1").replace(",", "")) or 1.0
            except (ValueError, TypeError):
                units = 1.0

            try:
                amount = float(str(row.get("amount") or "0").replace(",", ""))
            except (ValueError, TypeError):
                amount = 0.0

            model = (row.get("model") or "").strip() or (row.get("items") or "").strip() or None
            variant = (row.get("variant") or "").strip() or None

            status_raw = (row.get("reconciliation_status") or "").strip().lower()
            reconciliation_status = "Reconciled" if status_raw.startswith("rec") else "Pending"

            # REPORTS-DEPT-1 Phase 4c: Opsra-user matching still runs
            # quietly in the background (recorded_by/source_team, for
            # future team/department attribution) — but rep_name_raw, the
            # literal sheet text, is what filtering and display now use.
            # A rep with no matching Opsra user still gets a fully usable
            # row; nothing here depends on the match succeeding.
            rep_id, rep_team = _get_rep(rep_name_raw)

            if wm_date:
                try:
                    row_date = datetime.strptime(sale_date, "%Y-%m-%d").date()
                    if row_date <= wm_date:
                        already_imported.append({"row": row_num, "region": region, "sale_date": sale_date})
                except Exception:
                    pass

            # Duplicate key now uses the raw rep name text, not the
            # matched Opsra user id — two different real reps who both
            # failed to match (recorded_by=None) would previously have
            # been falsely treated as potential duplicates of each other.
            dup_key = ((rep_name_raw or "").lower(), sale_date, (customer_name or "").lower(), (model or "").lower(), amount)
            if dup_key in existing_keys:
                duplicate_warnings.append({"row": row_num, "region": region, "sale_date": sale_date, "customer_name": customer_name})

            notes = None
            if rep_name_raw and not rep_id:
                notes = f"No matching Opsra user for this rep — attribution unavailable, but the row is otherwise complete."

            valid_rows.append({
                "org_id":                org_id,
                "customer_name":         customer_name,
                "amount":                amount,
                "currency":              "NGN",
                "sale_date":             sale_date,
                "channel":               "other",
                "region":                region,
                "model":                 model,
                "variant":               variant,
                "units":                 units,
                "reconciliation_status": reconciliation_status,
                "rep_name":              rep_name_raw,
                "source_team":           rep_team,
                "recorded_by":           rep_id,
                "notes":                 notes,
                "import_source":         import_source,
                "created_at":            now_iso,
                "updated_at":            now_iso,
            })

    return {
        "valid_rows":         valid_rows,
        "error_rows":         error_rows,
        "duplicate_warnings": duplicate_warnings,
        "already_imported":   already_imported,
    }


def _parse_date(value: str) -> Optional[str]:
    if not value:
        return None
    v = str(value).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(v, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _rows_to_dicts(rows: list) -> list[dict]:
    if not rows:
        raise ValueError("File is empty — no rows found.")
    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_map: dict[int, str] = {}
    for idx, raw_h in enumerate(raw_headers):
        canonical = _normalise_header(raw_h)
        if canonical:
            col_map[idx] = canonical
    if not col_map:
        raise ValueError(
            "No recognised columns found. Expected at least one of: "
            + ", ".join(sorted(CANONICAL_FIELDS))
        )
    result: list[dict] = []
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record: dict = {field: None for field in CANONICAL_FIELDS}
        for idx, canonical in col_map.items():
            cell_val = row[idx] if idx < len(row) else None
            record[canonical] = str(cell_val).strip() if cell_val is not None else None
        result.append(record)
    return result


def _parse_xlsx(file_bytes: bytes) -> list:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    return list(wb.active.iter_rows(values_only=True))


def _parse_xls(file_bytes: bytes) -> list:
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError(
            "Legacy .xls files require xlrd. "
            "Run: pip install xlrd --break-system-packages"
        ) from exc
    wb = xlrd.open_workbook(file_contents=file_bytes)
    ws = wb.sheet_by_index(0)
    rows = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            if cell.ctype == 3:
                try:
                    dt = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                    row.append(datetime(*dt[:3]).strftime("%Y-%m-%d") if dt[0] else "")
                except Exception:
                    row.append(str(cell.value))
            else:
                row.append(cell.value)
        rows.append(tuple(row))
    return rows


def _parse_csv(file_bytes: bytes) -> list:
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1")
    return [tuple(r) for r in csv.reader(io.StringIO(text))]


# ---------------------------------------------------------------------------
# Public: parse file
# ---------------------------------------------------------------------------

def parse_excel_file(file_bytes: bytes, row_mapper=None) -> list[dict]:
    """
    Parse xlsx, xls, or csv bytes. Uses magic-byte detection.
    row_mapper: optional callable(rows) -> list[dict]. Defaults to
    _rows_to_dicts (named-customer aliases). Pass _aggregate_rows_to_dicts
    for daily-aggregate sales sheets.
    S14: raises ValueError on any parse failure.
    """
    if len(file_bytes) < 4:
        raise ValueError("File is too small to be a valid spreadsheet.")
    magic = file_bytes[:4]
    try:
        if magic[:4] == _XLSX_MAGIC or magic[:2] == b'PK':
            rows = _parse_xlsx(file_bytes)
        elif magic == _XLS_MAGIC:
            rows = _parse_xls(file_bytes)
        else:
            rows = _parse_csv(file_bytes)
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"Could not read file: {exc}") from exc
    mapper = row_mapper or _rows_to_dicts
    return mapper(rows)


# ---------------------------------------------------------------------------
# Public: fetch Google Sheet
# ---------------------------------------------------------------------------

def fetch_sheets_raw_rows(url: str) -> list[dict]:
    """
    Fetch a publicly shared Google Sheet as CSV, returning rows keyed by
    their ORIGINAL header text, unmapped. Split out of fetch_sheets_csv()
    so a caller can apply a different column-alias set (e.g. the
    daily-aggregate importer) without duplicating the network/CSV logic.
    S14: raises ValueError on any failure.
    """
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not match:
        raise ValueError(
            "Invalid Google Sheets URL. Expected: "
            "https://docs.google.com/spreadsheets/d/{id}/..."
        )
    export_url = (
        f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=csv"
    )
    try:
        response = httpx.get(export_url, timeout=15.0, follow_redirects=True)
    except httpx.RequestError as exc:
        raise ValueError(f"Network error fetching Google Sheet: {exc}") from exc
    except Exception as exc:
        raise ValueError(f"Error fetching Google Sheet: {exc}") from exc

    if response.status_code != 200:
        raise ValueError(
            f"Could not fetch Google Sheet (HTTP {response.status_code}). "
            "Make sure the sheet is set to 'Anyone with link can view'."
        )
    try:
        raw_rows = list(csv.DictReader(io.StringIO(response.text)))
    except Exception as exc:
        raise ValueError(f"Could not parse sheet as CSV: {exc}") from exc

    if not raw_rows:
        raise ValueError("Google Sheet is empty — no data rows found.")
    return raw_rows


def fetch_sheets_csv(url: str) -> list[dict]:
    """
    Fetch a publicly shared Google Sheet as CSV, mapped via the
    named-customer alias set. Unchanged public behaviour — now
    implemented on top of fetch_sheets_raw_rows().
    """
    raw_rows = fetch_sheets_raw_rows(url)
    result: list[dict] = []
    for raw_row in raw_rows:
        record: dict = {field: None for field in CANONICAL_FIELDS}
        for raw_key, cell_val in raw_row.items():
            canonical = _normalise_header(raw_key or "")
            if canonical:
                record[canonical] = str(cell_val).strip() if cell_val else None
        result.append(record)
    return result


def fetch_aggregate_sheets_csv(url: str) -> list[dict]:
    """
    Same fetch as fetch_sheets_csv(), mapped via _AGGREGATE_ALIASES
    instead, for daily aggregate sales sheets.
    """
    raw_rows = fetch_sheets_raw_rows(url)
    result: list[dict] = []
    for raw_row in raw_rows:
        record: dict = {field: None for field in set(_AGGREGATE_ALIASES.values())}
        for raw_key, cell_val in raw_row.items():
            canonical = _normalise_header(raw_key or "", _AGGREGATE_ALIASES)
            if canonical:
                record[canonical] = str(cell_val).strip() if cell_val else None
        result.append(record)
    return result


# ---------------------------------------------------------------------------
# Public: watermark helpers
# ---------------------------------------------------------------------------

def get_watermark(db, org_id: str, source_type: str, sheet_url: Optional[str]) -> Optional[str]:
    """
    Return the last_imported_date (ISO string YYYY-MM-DD) for this org+source,
    or None if no watermark exists yet.
    sheet_url is None for excel imports.
    S14: returns None on any DB failure.
    """
    try:
        query = (
            db.table("sales_import_watermarks")
            .select("last_imported_date")
            .eq("org_id", org_id)
            .eq("source_type", source_type)
        )
        if sheet_url:
            query = query.eq("sheet_url", sheet_url)
        else:
            query = query.is_("sheet_url", "null")
        res = query.maybe_single().execute()
        data = res.data
        if isinstance(data, list):
            data = data[0] if data else None
        if data and data.get("last_imported_date"):
            return str(data["last_imported_date"])[:10]
    except Exception:
        logger.warning("GPM-1E: get_watermark query failed — treating as no watermark.")
    return None


def save_watermark(
    db,
    org_id: str,
    source_type: str,
    sheet_url: Optional[str],
    last_date: str,
) -> None:
    """
    Upsert the watermark for this org+source to last_date (YYYY-MM-DD).
    S14: logs and swallows any DB failure — import already succeeded.
    """
    from datetime import timezone
    try:
        db.table("sales_import_watermarks").upsert(
            {
                "org_id":              org_id,
                "source_type":         source_type,
                "sheet_url":           sheet_url,
                "last_imported_date":  last_date,
                "last_imported_at":    datetime.now(timezone.utc).isoformat(),
            },
            on_conflict="org_id,source_type,sheet_url",
        ).execute()
    except Exception:
        logger.warning("GPM-1E: save_watermark failed — watermark not updated.")


def reset_watermark(db, org_id: str, source_type: str, sheet_url: Optional[str]) -> None:
    """
    Delete the watermark for this org+source so the next import starts from scratch.
    S14: swallows any failure.
    """
    try:
        query = (
            db.table("sales_import_watermarks")
            .delete()
            .eq("org_id", org_id)
            .eq("source_type", source_type)
        )
        if sheet_url:
            query = query.eq("sheet_url", sheet_url)
        else:
            query = query.is_("sheet_url", "null")
        query.execute()
    except Exception:
        logger.warning("GPM-1E: reset_watermark failed.")


# ---------------------------------------------------------------------------
# Public: validate rows
# ---------------------------------------------------------------------------

def validate_and_prepare_rows(
    rows: list[dict],
    org_id: str,
    db,
    import_source: str,
    watermark_date: Optional[str] = None,
) -> dict:
    """
    Validate raw rows and prepare them for insertion into direct_sales.

    Returns:
      {
        valid_rows:          list[dict]              — ready to INSERT (all non-error rows)
        error_rows:          list[{row, message}]    — validation failures
        duplicate_warnings:  list[{row, phone, sale_date, amount}] — triple-match duplicates
        already_imported:    list[{row, sale_date}]  — rows at/before watermark date
      }

    Rows in duplicate_warnings and already_imported are STILL in valid_rows.
    The route + frontend use selected_indices to decide which rows to actually insert.

    Duplicate check: phone + sale_date + amount all match an existing direct_sales row.
    Watermark check: sale_date <= watermark_date → flagged as already_imported.
    """
    from datetime import timezone
    from datetime import timezone
    now_iso = datetime.now(timezone.utc).isoformat()

    valid_rows:         list[dict] = []
    error_rows:         list[dict] = []
    duplicate_warnings: list[dict] = []
    already_imported:   list[dict] = []

    # Parse watermark date once
    wm_date: Optional[date] = None
    if watermark_date:
        try:
            wm_date = datetime.strptime(watermark_date[:10], "%Y-%m-%d").date()
        except Exception:
            pass

    # Load existing (phone, sale_date, amount) triples for duplicate detection
    existing_triples: set[tuple] = set()
    try:
        res = (
            db.table("direct_sales")
            .select("phone,sale_date,amount")
            .eq("org_id", org_id)
            .execute()
        )
        for r in (res.data or []):
            if r.get("phone") and r.get("sale_date") and r.get("amount") is not None:
                existing_triples.add((
                    str(r["phone"]).strip(),
                    str(r["sale_date"])[:10],
                    float(r["amount"]),
                ))
    except Exception:
        logger.warning("GPM-1E: Duplicate check query failed — skipping.")

    for i, row in enumerate(rows):
        row_num = i + 2  # 1-indexed; row 1 = header

        # --- customer_name required ---
        customer_name = (row.get("customer_name") or "").strip()
        if not customer_name:
            error_rows.append({"row": row_num, "message": "customer_name is required"})
            continue

        # --- amount required, numeric > 0 ---
        raw_amount = str(row.get("amount") or "").strip()
        try:
            amount = float(raw_amount.replace(",", ""))
            if amount <= 0:
                raise ValueError()
        except (ValueError, TypeError):
            error_rows.append({"row": row_num, "message": f"Invalid amount: '{raw_amount}'"})
            continue

        # --- sale_date required, parseable ---
        raw_date = str(row.get("sale_date") or "").strip()
        sale_date = _parse_date(raw_date)
        if not sale_date:
            error_rows.append({
                "row": row_num,
                "message": (
                    f"Invalid sale_date: '{raw_date}'. "
                    "Use YYYY-MM-DD, DD/MM/YYYY, or MM/DD/YYYY"
                ),
            })
            continue

        # --- optional fields ---
        phone       = (row.get("phone")       or "").strip() or None
        region      = (row.get("region")      or "").strip() or None
        channel     = (row.get("channel")     or "").strip() or "other"
        source_team = (row.get("source_team") or "").strip() or None
        notes       = (row.get("notes")       or "").strip() or None

        # --- watermark check ---
        if wm_date:
            try:
                row_date = datetime.strptime(sale_date, "%Y-%m-%d").date()
                if row_date <= wm_date:
                    already_imported.append({
                        "row":       row_num,
                        "sale_date": sale_date,
                        "customer":  customer_name,
                    })
            except Exception:
                pass

        # --- duplicate check (phone + sale_date + amount) ---
        if phone and (phone, sale_date, amount) in existing_triples:
            duplicate_warnings.append({
                "row":       row_num,
                "phone":     phone,
                "sale_date": sale_date,
                "amount":    amount,
            })

        valid_rows.append({
            "org_id":        org_id,
            "customer_name": customer_name,
            "amount":        amount,
            "currency":      "NGN",
            "sale_date":     sale_date,
            "channel":       channel,
            "phone":         phone,
            "region":        region,
            "source_team":   source_team,
            "notes":         notes,
            "import_source": import_source,
            "created_at":    now_iso,
            "updated_at":    now_iso,
        })

    return {
        "valid_rows":         valid_rows,
        "error_rows":         error_rows,
        "duplicate_warnings": duplicate_warnings,
        "already_imported":   already_imported,
    }


# ---------------------------------------------------------------------------
# REPORTS-DEPT-1 Phase 4: validate daily-aggregate rows
# ---------------------------------------------------------------------------

def validate_and_prepare_aggregate_sales_rows(
    rows: list[dict],
    org_id: str,
    db,
    import_source: str,
    watermark_date: Optional[str] = None,
) -> dict:
    """
    Validate and prepare rows from a daily AGGREGATE sales sheet (one row
    per day per rep, with separate per-product-line revenue columns) for
    insertion into direct_sales.

    Unlike validate_and_prepare_rows() (named-customer transactions):
      - customer_name is NOT required (confirmed nullable at the DB level
        before building this — no placeholder value invented).
      - each sheet row becomes ONE direct_sales row PER product line with
        revenue > 0, tagged with product_line, so Revenue stays genuinely
        filterable by line rather than one opaque summed total.
      - rep_name is resolved against users.full_name for this org via
        _match_rep(); a match populates recorded_by + source_team (same
        attribution mechanism as REPORTS-DEPT-1 Phase 0 — not a new one).
        An unmatched name doesn't block the row, it just imports without
        attribution and a note recording what was on the sheet.

    Returns the same shape as validate_and_prepare_rows().
    """
    from datetime import timezone
    now_iso = datetime.now(timezone.utc).isoformat()

    valid_rows:         list[dict] = []
    error_rows:         list[dict] = []
    duplicate_warnings: list[dict] = []
    already_imported:   list[dict] = []

    wm_date: Optional[date] = None
    if watermark_date:
        try:
            wm_date = datetime.strptime(watermark_date[:10], "%Y-%m-%d").date()
        except Exception:
            pass

    # Duplicate check keyed on (rep, sale_date, product_line, amount) —
    # no phone/customer to key on for aggregate rows.
    existing_keys: set = set()
    try:
        res = (
            db.table("direct_sales")
            .select("recorded_by,sale_date,product_line,amount")
            .eq("org_id", org_id)
            .eq("import_source", import_source)
            .execute()
        )
        for r in (res.data or []):
            if r.get("sale_date") and r.get("amount") is not None:
                existing_keys.add((
                    r.get("recorded_by"),
                    str(r["sale_date"])[:10],
                    r.get("product_line"),
                    float(r["amount"]),
                ))
    except Exception:
        logger.warning("Phase 4 aggregate import: duplicate check query failed — skipping.")

    # Cache rep matches within this run — one rep typically appears on
    # many rows (many days), so avoid one DB query per row.
    rep_cache: dict = {}

    def _get_rep(rep_name: Optional[str]):
        if not rep_name:
            return None, None
        key = rep_name.strip().lower()
        if key not in rep_cache:
            rep_cache[key] = _match_rep(rep_name, org_id, db)
        return rep_cache[key]

    PRODUCT_LINES = [
        ("mattress_revenue", "mattress"),
        ("pillow_revenue",   "pillow"),
    ]

    for i, row in enumerate(rows):
        row_num = i + 2  # 1-indexed; row 1 = header

        raw_date = str(row.get("sale_date") or "").strip()
        sale_date = _parse_date(raw_date)
        if not sale_date:
            error_rows.append({
                "row": row_num,
                "message": (
                    f"Invalid sale_date: '{raw_date}'. "
                    "Use YYYY-MM-DD, DD/MM/YYYY, or MM/DD/YYYY"
                ),
            })
            continue

        rep_name_raw = (row.get("rep_name") or "").strip() or None
        rep_id, rep_team = _get_rep(rep_name_raw)

        row_has_revenue = False
        for field, product_line in PRODUCT_LINES:
            raw_amount = str(row.get(field) or "").strip()
            if not raw_amount:
                continue
            try:
                amount = float(raw_amount.replace(",", ""))
            except (ValueError, TypeError):
                error_rows.append({
                    "row": row_num,
                    "message": f"Invalid {field.replace('_', ' ')}: '{raw_amount}'",
                })
                continue
            if amount <= 0:
                continue  # blank/zero for this line — not an error, nothing to import

            row_has_revenue = True

            if wm_date:
                try:
                    row_date = datetime.strptime(sale_date, "%Y-%m-%d").date()
                    if row_date <= wm_date:
                        already_imported.append({
                            "row": row_num, "sale_date": sale_date,
                            "product_line": product_line,
                        })
                except Exception:
                    pass

            dup_key = (rep_id, sale_date, product_line, amount)
            if dup_key in existing_keys:
                duplicate_warnings.append({
                    "row": row_num, "sale_date": sale_date,
                    "product_line": product_line, "amount": amount,
                })

            notes = None
            if rep_name_raw and not rep_id:
                notes = f"Unmatched rep on sheet: {rep_name_raw}"

            valid_rows.append({
                "org_id":        org_id,
                "customer_name": None,
                "amount":        amount,
                "currency":      "NGN",
                "sale_date":     sale_date,
                "channel":       "other",
                "product_line":  product_line,
                "source_team":   rep_team,
                "recorded_by":   rep_id,
                "notes":         notes,
                "import_source": import_source,
                "created_at":    now_iso,
                "updated_at":    now_iso,
            })

        if not row_has_revenue:
            error_rows.append({
                "row": row_num,
                "message": "No revenue found for this row (Mattress Revenue and Pillow Revenue both blank/zero).",
            })

    return {
        "valid_rows":         valid_rows,
        "error_rows":         error_rows,
        "duplicate_warnings": duplicate_warnings,
        "already_imported":   already_imported,
    }
